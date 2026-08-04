"""
AMECATH Executive Dashboards — 7-Question Summary + Master Consolidated Dashboard
-----------------------------------------------------------------------------------
1. Opens each country workbook.
2. Appends a new "7_Questions_Summary" sheet answering the 7 standard market
   questions, styled to match each workbook's existing conventions
   (Arial font, navy header 1F3864, blue sub-header 2E5395, banded rows D9E2F3).
3. Builds one Master Consolidated Dashboard workbook that:
     - Includes a "Master_7_Questions" sheet with all 8 countries side by side.
     - Copies every existing sheet from every country file into the master
       workbook, tagged with Source Country + Original Sheet tracking columns.

Run:  python3 build_dashboards.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import os

SRC_DIR = "/home/claude/work"
OUT_DIR = "/mnt/user-data/outputs"
os.makedirs(OUT_DIR, exist_ok=True)

FONT_NAME = "Arial"
NAVY = "1F3864"
BLUE = "2E5395"
BAND = "D9E2F3"
WHITE = "FFFFFF"

TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name=FONT_NAME, size=11, color="595959")
SECTION_FONT = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
BODY_FONT = Font(name=FONT_NAME, size=10, color="000000")

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=BLUE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BAND_FILL = PatternFill("solid", fgColor=BAND)
NO_FILL = PatternFill(fill_type=None)

THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_TOP_BOLD = Alignment(wrap_text=True, vertical="top", horizontal="left")

# ---------------------------------------------------------------------------
# Source content — condensed from the AMECATH regional research brief.
# Each entry maps directly to the 7 required questions, in order.
# ---------------------------------------------------------------------------

QUESTIONS = [
    "1. % Increase per Year (2021-2027) & 2027 Prediction",
    "2. Number of Chronic Hemodialysis Patients",
    "3. Number of Nephrologists",
    "4. Number of Vascular Surgeons",
    "5. Number of Hemodialysis Centers / Units / Hospitals",
    "6. Number of Hemodialysis Machines",
    "7. Number of HD Catheters Consumed in 2025",
]

COUNTRY_DATA = {
    "Saudi_Arabia": {
        "file": "AMECATH_Saudi_Arabia_Executive_Dashboard.xlsx",
        "answers": [
            "~9% YoY consistent annual growth in ESRD/dialysis population; total dialysis patients projected to reach 30,000-33,000 by 2027.",
            "20,534 chronic hemodialysis patients (Saudi Center for Organ Transplantation - SCOT registries).",
            "1,279 nephrologists, ~39 per million population.",
            "No standalone national registry; AV fistula/tunneled catheter creation handled by specialized vascular surgeons and high-volume general surgeons across MOH clusters, Diaverum centers, and private tertiary networks (e.g., Saudi German Health).",
            "270+ official dialysis clinics/centers nationwide (Diaverum alone operates 35 clinics across 27 cities).",
            "Estimated 8,000+ machines integrated across MOH, military, and private sectors.",
            "Volume driven by central NUPCO tenders and private networks; AMECATH targets 3%-5% market share of addressable catheter/vascular-access volume.",
        ],
        "source": "SCOT registries; NUPCO tenders; Diaverum footprint reporting; AMECATH internal targets",
    },
    "UAE": {
        "file": "AMECATH_UAE_Executive_Dashboard.xlsx",
        "answers": [
            "Estimated 5%-6% CAGR driven by rapid population growth and high diabetes prevalence, with projections scaling upward through 2027.",
            "Fragmented across emirate-level providers (SEHA, DHA, DoH); thousands of active patients managed across integrated networks (e.g., new SEHA facilities accommodating 100+ baseline patients per rollout).",
            "~14-20+ per million population across major hubs (regulatory licensing baseline).",
            "Concentrated in high-end referral hubs: Tawam Hospital, Sheikh Khalifa Medical City (SKMC), and private groups (Mediclinic, NMC, Burjeel).",
            "Dozens of specialized facilities, anchored by SEHA Kidney Care (multiple centers across Abu Dhabi, Al Ain, Al Dhafra) and DHA-regulated Dubai providers.",
            "Hundreds of advanced machines distributed across public health clusters and private tertiary hospitals.",
            "UAE hemodialysis catheter market commercially valued at approximately USD 14 million.",
        ],
        "source": "SEHA / DHA / DoH regulatory data; hospital network reporting; market valuation estimate",
    },
    "Qatar": {
        "file": "AMECATH_Qatar_Executive_Dashboard.xlsx",
        "answers": [
            "Average annual increase of 5.67% (APC) from 2022 to 2030, per established clinical forecasting model.",
            "1,050 HD patients (plus 275 peritoneal dialysis patients; 1,325 prevalent renal patients total).",
            "Centralized entirely under Hamad Medical Corporation (HMC) nephrology leadership (no separate national count published).",
            "Integrated within HMC's centralized surgical and renal vascular teams at Hamad General Hospital and Fahad Bin Jassim Kidney Center.",
            "7 dedicated national HD centers, incl. flagship Fahad Bin Jassim Kidney Center (92 stations) plus regional units at Al Khor, Al Wakra, and Hazm Mebaireek.",
            "197 HD stations/machines distributed across the 7 ambulatory and hospital-based units.",
            "Procured centrally via HMC's unified national purchasing gateway, matching the 5.67% annual patient growth projection.",
        ],
        "source": "Hamad Medical Corporation (HMC) clinical forecasting model and procurement gateway",
    },
    "Kuwait": {
        "file": "AMECATH_Kuwait_Executive_Dashboard.xlsx",
        "answers": [
            "Stable baseline (~0.047% of total general population) with steady incremental yearly growth; no single published 2027 figure.",
            "2,275 total dialysis patients (HD ~89% share).",
            "Estimated 21-49 per million population, coordinated via the Kuwaiti Association of Nephrology (KNA).",
            "Primarily centralized in major public renal blocks: Mohammed Al-Khazam Kidney Center (Al-Adan Hospital), Amiri, and Jahra hospitals.",
            "87 dialysis units across major public kidney centers and specialized expansions (e.g., Sabah Al-Ahmad Health Center unit: 20 machines / 120-patient capacity), plus select private facilities (New Mowasat, Salam Hospital).",
            "932 machines registered nationally under the Ministry of Health infrastructure database.",
            "Regulated via Central Tenders Committee (CTC) public procurement and private hospital reorder streams.",
        ],
        "source": "Kuwait MOH infrastructure database; Kuwaiti Association of Nephrology (KNA); Central Tenders Committee (CTC)",
    },
    "Oman": {
        "file": "AMECATH_Oman_Executive_Dashboard.xlsx",
        "answers": [
            "Steady expansion mirroring regional trends (~5% annual growth), driven by diabetic nephropathy and aging demographics.",
            "2,500 to 4,500 active patients managed across the Ministry of Health network.",
            "Average density estimated around 32-40 per million population.",
            "Centered predominantly at the Royal Hospital and major regional referral hospitals (North Batinah, Salalah, etc.).",
            "24 core MOH renal dialysis centers and expansions (Renal Dialysis Center opposite the Royal Hospital, Bousher, Al Amerat, plus new regional setups in Sur and Al Khaboura).",
            "Fully equipped across 24+ regional and capital-governorate public renal units (no separate machine count published).",
            "Procured primarily via Oman's Ministry of Health centralized tenders; AMECATH targets 4%-6% market share for specialized vascular access portfolios.",
        ],
        "source": "Oman Ministry of Health centralized tenders and regional renal unit reporting",
    },
    "Bahrain": {
        "file": "AMECATH_Bahrain_Executive_Dashboard.xlsx",
        "answers": [
            "Highly regulated and tracked via the National Health Regulatory Authority (NHRA); stable, predictable yearly increments (no single published % figure).",
            "4,547 active dialysis patients (4,298 Bahraini nationals + 249 non-Bahraini residents), generating 47,000+ annual treatment sessions.",
            "Specialized clinical staff covering Salmaniya Medical Complex and BDF Royal Medical Services (no separate national count published).",
            "Operating across tertiary public setups and leading private hospitals (no separate national count published).",
            "Key public sites: Salmaniya Medical Complex and Abdulrahman Kanoo Dialysis Centre (Muharraq); private providers with dedicated dialysis beds (Bahrain Specialist Hospital: 8 beds; Royal Bahrain Hospital / KIMSHEALTH: 10 units).",
            "Distributed across public renal centers and private hospital inpatient/outpatient units (no separate national machine count published).",
            "Compact, highly transparent market; AMECATH targets 5%-8% market share in high-volume accounts.",
        ],
        "source": "National Health Regulatory Authority (NHRA); Salmaniya Medical Complex; AMECATH internal targets",
    },
    "Jordan": {
        "file": "AMECATH_Jordan_Executive_Dashboard.xlsx",
        "answers": [
            "Predictable 4%-6% annual growth, per the Jordan MOH ESRD Registry.",
            "6,063 active ESRD patients (5,960 on HD, 103 on PD).",
            "Robust academic and clinical presence backed by Jordan University Hospital and King Abdullah University Hospital (no single national headcount published).",
            "Widely available across Amman's advanced private hospital network (Al-Khalidi, Islamic Hospital, etc.) and public teaching hospitals.",
            "86 total units nationwide: MOH (37), Private Sector (36), Royal Medical Services / Military, and University Hospitals.",
            "932 machines nationally, by sector: MOH (435), Private Sector (379), RMS Military (84), Academic Universities (34).",
            "High consumption volume driven by competitive tender dynamics between public institutions and private tertiary centers.",
        ],
        "source": "Jordan MOH ESRD Registry",
    },
    "Lebanon": {
        "file": "AMECATH_Lebanon_Executive_Dashboard.xlsx",
        "answers": [
            "Despite macroeconomic headwinds, clinical demand remains structurally rigid; no single published % growth figure, thousands of maintenance sessions required weekly.",
            "~4,400 prevalent patients (~4,190 on HD, ~95% share; ~210 on PD).",
            "189 registered nephrologists under the Lebanese Society of Nephrology and Hypertension (LSNH), ~36 per million.",
            "Concentrated heavily in Beirut and Mount Lebanon university medical centers (AUBMC, Hotel-Dieu de France, Saint George Hospital).",
            "80 hospital-based dialysis units nationwide (14 public, 66 private) — Lebanese regulations mandate dialysis be hospital-affiliated.",
            "Distributed across the 80 certified hospital-based renal units (no separate national machine count published).",
            "Supplied via private medical importers and hospital-supply distributors; AMECATH targets 3%-5% initial vascular-access volume share.",
        ],
        "source": "Lebanese Society of Nephrology and Hypertension (LSNH)",
    },
}


def style_summary_sheet(ws, country_label, answers, source_note):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 90

    # Title
    ws.merge_cells("A1:B1")
    ws["A1"] = "7 Questions Summary"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:B2")
    ws["A2"] = f"AMECATH — {country_label} Renal Access Market Intelligence"
    ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 18

    # Section header
    ws.merge_cells("A4:B4")
    ws["A4"] = "Standard Market Questions (1-7)"
    ws["A4"].font = SECTION_FONT
    ws["A4"].fill = SECTION_FILL
    ws.row_dimensions[4].height = 20

    # Table header
    ws["A5"] = "Metric"
    ws["B5"] = "Details"
    for c in ("A5", "B5"):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL
        ws[c].alignment = Alignment(vertical="center")
        ws[c].border = BORDER
    ws.row_dimensions[5].height = 18

    row = 6
    for i, (q, a) in enumerate(zip(QUESTIONS, answers)):
        cell_a = ws.cell(row=row, column=1, value=q)
        cell_b = ws.cell(row=row, column=2, value=a)
        fill = BAND_FILL if i % 2 == 1 else NO_FILL
        for cell in (cell_a, cell_b):
            cell.font = BODY_FONT
            cell.fill = fill
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        ws.row_dimensions[row].height = 44
        row += 1

    # Source note
    row += 1
    ws.merge_cells(f"A{row}:B{row}")
    src_cell = ws.cell(row=row, column=1, value=f"Primary source references: {source_note}")
    src_cell.font = Font(name=FONT_NAME, size=9, italic=True, color="595959")
    src_cell.alignment = WRAP_TOP
    ws.row_dimensions[row].height = 28


def add_7q_sheet_to_country_file(country_key, info):
    path = os.path.join(SRC_DIR, info["file"])
    wb = openpyxl.load_workbook(path)

    sheet_name = "7_Questions_Summary"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    country_label = country_key.replace("_", " ")
    style_summary_sheet(ws, country_label, info["answers"], info["source"])

    # Place the new sheet right after the first sheet for visibility
    wb.move_sheet(sheet_name, offset=-(len(wb.sheetnames) - 2))

    out_path = os.path.join(OUT_DIR, info["file"])
    wb.save(out_path)
    print(f"Updated: {out_path}")
    return out_path


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def build_master_workbook(country_files):
    master = openpyxl.Workbook()
    default_sheet = master.active
    master.remove(default_sheet)

    # --- Master 7-Questions comparison sheet -------------------------------
    ws = master.create_sheet("Master_7_Questions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 44
    for idx in range(len(COUNTRY_DATA)):
        ws.column_dimensions[get_column_letter(idx + 2)].width = 42

    ws.merge_cells(f"A1:{get_column_letter(len(COUNTRY_DATA)+1)}1")
    ws["A1"] = "Master Consolidated Dashboard — 7 Questions Summary (All Countries)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws.row_dimensions[1].height = 22

    ws["A3"] = "Metric"
    ws["A3"].font = HEADER_FONT
    ws["A3"].fill = HEADER_FILL
    ws["A3"].border = BORDER

    countries_sorted = sorted(COUNTRY_DATA.keys())
    for i, ck in enumerate(countries_sorted):
        cell = ws.cell(row=3, column=i + 2, value=ck.replace("_", " "))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[3].height = 18

    for qi, q in enumerate(QUESTIONS):
        r = 4 + qi
        a_cell = ws.cell(row=r, column=1, value=q)
        fill = BAND_FILL if qi % 2 == 1 else NO_FILL
        a_cell.font = Font(name=FONT_NAME, size=10, bold=True)
        a_cell.fill = fill
        a_cell.alignment = WRAP_TOP
        a_cell.border = BORDER
        for ci, ck in enumerate(countries_sorted):
            val = COUNTRY_DATA[ck]["answers"][qi]
            c = ws.cell(row=r, column=ci + 2, value=val)
            c.font = BODY_FONT
            c.fill = fill
            c.alignment = WRAP_TOP
            c.border = BORDER
        ws.row_dimensions[r].height = 60

    # --- Consolidated Data Index sheet (tracking) ---------------------------
    idx_ws = master.create_sheet("Consolidated_Data_Index")
    idx_ws.sheet_view.showGridLines = False
    headers = ["Source Country", "Original Sheet Name", "Master Sheet Name", "Row Count", "Column Count"]
    for ci, h in enumerate(headers):
        c = idx_ws.cell(row=1, column=ci + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    for ci, w in enumerate([20, 32, 40, 12, 14]):
        idx_ws.column_dimensions[get_column_letter(ci + 1)].width = w
    idx_ws.row_dimensions[1].height = 18

    index_row = 2
    used_names = set(master.sheetnames)

    # --- Copy every sheet from every country file into the master ----------
    for ck in countries_sorted:
        info = COUNTRY_DATA[ck]
        path = os.path.join(OUT_DIR, info["file"])
        src_wb = openpyxl.load_workbook(path, data_only=False)
        country_label = ck.replace("_", " ")

        for sheet_name in src_wb.sheetnames:
            src_ws = src_wb[sheet_name]

            # Build a safe, unique master sheet name (<=31 chars for Excel)
            short_country = "".join([w[:3] for w in country_label.split()])[:6]
            short_sheet = sheet_name.split(".", 1)[-1].strip() if "." in sheet_name else sheet_name
            base_name = f"{short_country}_{short_sheet}"
            base_name = base_name[:31]
            new_name = base_name
            suffix = 1
            while new_name in used_names:
                suffix += 1
                new_name = (base_name[: 31 - len(str(suffix)) - 1] + f"_{suffix}")
            used_names.add(new_name)

            dst_ws = master.create_sheet(new_name)

            # Tracking columns in first two data columns of every copied sheet
            dst_ws["A1"] = "Source Country"
            dst_ws["B1"] = "Original Sheet"
            dst_ws["A1"].font = HEADER_FONT
            dst_ws["B1"].font = HEADER_FONT
            dst_ws["A1"].fill = HEADER_FILL
            dst_ws["B1"].fill = HEADER_FILL
            dst_ws["A2"] = country_label
            dst_ws["B2"] = sheet_name
            dst_ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True)
            dst_ws["B2"].font = Font(name=FONT_NAME, size=10, italic=True)

            # Offset original content 2 rows down, 0 columns right, preserving style
            max_row = src_ws.max_row
            max_col = src_ws.max_column
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    src_cell = src_ws.cell(row=r, column=c)
                    if src_cell.value is None and not src_cell.has_style:
                        continue
                    dst_cell = dst_ws.cell(row=r + 3, column=c, value=src_cell.value)
                    copy_cell_style(src_cell, dst_cell)

            # Copy column widths
            for col_letter, dim in src_ws.column_dimensions.items():
                if dim.width:
                    dst_ws.column_dimensions[col_letter].width = dim.width
            dst_ws.column_dimensions["A"].width = max(dst_ws.column_dimensions["A"].width or 0, 20)
            dst_ws.column_dimensions["B"].width = max(dst_ws.column_dimensions["B"].width or 0, 32)

            # Copy merged cells (shifted down 3 rows)
            for merged_range in list(src_ws.merged_cells.ranges):
                min_col, min_row, max_col_m, max_row_m = merged_range.bounds
                dst_ws.merge_cells(
                    start_row=min_row + 3, start_column=min_col,
                    end_row=max_row_m + 3, end_column=max_col_m,
                )

            # Index row
            idx_ws.cell(row=index_row, column=1, value=country_label)
            idx_ws.cell(row=index_row, column=2, value=sheet_name)
            idx_ws.cell(row=index_row, column=3, value=new_name)
            idx_ws.cell(row=index_row, column=4, value=max_row)
            idx_ws.cell(row=index_row, column=5, value=max_col)
            for cc in range(1, 6):
                idx_ws.cell(row=index_row, column=cc).font = BODY_FONT
                idx_ws.cell(row=index_row, column=cc).border = BORDER
            index_row += 1

    # Reorder: Master_7_Questions, Consolidated_Data_Index first
    order = ["Master_7_Questions", "Consolidated_Data_Index"]
    order += [s for s in master.sheetnames if s not in order]
    master._sheets = [master[s] for s in order]

    out_path = os.path.join(OUT_DIR, "AMECATH_Master_Consolidated_Dashboard.xlsx")
    master.save(out_path)
    print(f"Created master workbook: {out_path}")
    return out_path


if __name__ == "__main__":
    updated_paths = []
    for ck, info in COUNTRY_DATA.items():
        updated_paths.append(add_7q_sheet_to_country_file(ck, info))

    master_path = build_master_workbook(updated_paths)
    print("\nDone.")
    print("Updated country files + master workbook are in:", OUT_DIR)
